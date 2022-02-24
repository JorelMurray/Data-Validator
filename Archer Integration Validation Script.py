#!/usr/bin/env python3

import os

import pandas as pd

import math

import xlrd

import numpy as np

import re

from pandas import ExcelWriter

from pandas import ExcelFile

from openpyxl import Workbook, load_workbook

from openpyxl.utils import get_column_letter

from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Fill, Color, colors

from openpyxl.styles.colors import Color

from openpyxl.cell import Cell


#Change directory to location of the excel workbooks

os.chdir("C:\\Users\\JMurray64\Documents\\Archer Integration Testing\\Integration Test 4")

nameExceptions = ["Kim, Soo Jung", "Carter, J Braxton", "Hua, Lindsay Jennie",  "Majji, Sree Mouli", "Udasco, Anthony Eric", "Lee, Hye Jung (Grace)", "Al Araj, Mohammad Haitham","Merla, Nageswara Rao", "Pappu, Lakshmi Sireesha", "Donthireddy, Venkata Reddy"]


#Function to remove ' (No Formatting)' text from headers in the Collibra SOX Report export

def removeFormatText(pathname):

    df = pd.read_excel(pathname)

    wb = load_workbook(pathname)

    ws = wb.active


    ws.title = "Sheet1"


    #Set colValues to the list of column names

    colValues = df.columns.values


    #Create to list arrays

    list1, list2 = [], []


    #Repeat for each column name

    for item in colValues:

        temp = []


        #If there is (No Formatting) in the column name, append it to list1 and a version without the (No Formatting) to list2

        if "(No Formatting)" in item:

            list1.append(item)

            temp = item.split(" (")

            list2.append(temp[0])


    #Create a dictionary matching list1 to list2

    newColumnNames = {list1[i]: list2[i] for i in range(len(list1))}


    #rRename the columns

    df.rename(columns = newColumnNames, inplace = True)

    df.to_excel(pathname, sheet_name = "Sheet1", index = False, header = True)


#Function to take the Archer SOX Report Export and parse ot the controls onto their own line

def parseSOXControls(pathname):

    wb = load_workbook(pathname)

    df = pd.read_excel(pathname)


    ws = wb.active


    reports = []

    rowIndex = 1


    #Populate array with a row for each report

    for row in ws.iter_rows(min_row = 2, max_col = 18, max_row = df.shape[0] + 1, values_only = True):

        if row[0] != None:

            reports.append(list(row))


        

    #Iterate through reports   

    for item in reports:


        #Proceed if the value in control column is not blank

        if item[9] != None:


            #split out controls

            tempControls = item[9].split("\n")

            numTempControls = len(tempControls)


        else:

            numTempControls = 0


        #Procede to next steps if there are more than one control for the report

        if numTempControls > 1:

            i = 1

            controlIndex = 0

            rowIndex += 1

            tempControls.sort()


            #Set the first report row to the 1st control

            ws.cell(row = rowIndex, column = 10).value = tempControls[controlIndex]


            #Create a loop of new row creation for each control

            while i < len(tempControls):

                rowIndex += 1


                #Insert a new row

                ws.insert_rows(rowIndex)


                #Loop through across the columns for the new row

                for r in range(0, 18):


                    #If it is the control column, insert the next control

                    if r == 9:

                        controlIndex = controlIndex + 1

                        ws.cell(row = rowIndex, column = 10).value = tempControls[controlIndex]

                    #Otherwise fill cells with the same data as above

                    else:

                        ws.cell(row = rowIndex, column = r + 1).value = item[r]


                i += 1

        #Pastes the row as is if there is only one control

        else:

            rowIndex += 1

            for r in range(0, 18):

                ws.cell(row = rowIndex, column = r + 1).value = item[r]

            

    wb.save(pathname)


def splitNames(columnName, df):

    #Loops through each item item in the dataset

    for i in df.index:


        if df[columnName][i] != "" and df[columnName][i] != None:


            #Splits each person into their own list item

            nameList = str(df[columnName][i]).split("\n")


            #Cycle through names to create a substring from start to second space

            for n in range(0, len(nameList)):


                if any(names in nameList[n] for names in nameExceptions):

                    nameMatch = [names in nameList[n] for names in nameExceptions]

                    nameIndex = nameMatch.index(True)

                    nameList[n] = nameExceptions[nameIndex]

                else:

                    #if nameExceptions in nameList[n]

                    end = nameList[n].find(",") + 1

                    end = nameList[n].find(" ", end + 1)  

                    nameList[n] = nameList[n][0:end]


            #Sort list and join to one string

            nameList.sort()   

            names = "; ".join(nameList)


            #Replace oringal entry with newly modified string

            df[columnName][i] = names


def replaceData(columnName, columns, targetColumnNum, ws, df):


    colNum = 0

    rowNum = 1


    #iterate over row in spreadsheet, rows & columns = max row/column of spreadhseet. Target Column is column to place the data

    for row in ws.iter_rows(min_row = 2, max_col = columns, max_row = df.shape[0] + 1, values_only = True):

        rowNum += 1

        if row[0] != None and row[0] != "" and not row[0] == xlrd.XL_CELL_BLANK:

            colNum = 0

            try:

                for item in row:

                    colNum += 1

                    if colNum == targetColumnNum:

                        ws.cell(row = rowNum, column = targetColumnNum).value = df[columnName][rowNum-2]

            except:

                pass   



def reformatNames(docName):


    #Executes on the Archer Controls document

    if docName == "Archer Controls":

        pathName = "TestArcherControlValidation.xlsx"


        df = pd.read_excel(pathName)

        wb = load_workbook(pathName)

        ws = wb.active


        df.replace(np.nan, '', regex=True, inplace = True)


        #Reformats the various columns

        splitNames("Functional Group Owner", df)

        splitNames("Point of Contact", df)

        splitNames("Control Owner", df)


        #Fills the spreadsheet with the newly formatted values

        replaceData("Functional Group Owner", 19, 15, ws, df)

        replaceData("Point of Contact", 19, 12, ws, df)

        replaceData("Control Owner", 19, 6, ws, df)


    #Executes on the Archer Business Process document

    if docName == "Archer Business Process":

        pathName = "TestArcherBusinessProcessValidation.xlsx"


        df = pd.read_excel(pathName)

        wb = load_workbook(pathName)

        ws = wb.active


        df.replace(np.nan, '', regex=True, inplace = True)


        #Remove string leading and trailing spaces

        df["Description"] = df["Description"].str.strip()


        #print(df.at[2,"Description"] == "Employee and Dealer Commission:  Commissions based on Sales and compensation plan for employees and dealer commission based on contract or SPIFF.  The process confirms approval of the plan as well as the technical execution of it and validates the transaction flow from beginning to end.")


        #Reformats the various columns

        splitNames("Business Process Owner", df)

        splitNames("Business Process Participants",df)


        #Fills the spreadsheet with the newly formatted values

        replaceData("Business Process Owner", 5, 4, ws, df)

        replaceData("Business Process Participants", 5, 5, ws, df)


    if docName == "Archer Reports":

        pathName = "TestArcherSOXReportValidation.xlsx"


        df = pd.read_excel(pathName)

        wb = load_workbook(pathName)

        ws = wb.active


        df.replace(np.nan, '', regex=True, inplace = True)


        #Reformats the various columns

        splitNames("Report Coordinator", df)


        #Fills the spreadsheet with the newly formatted values

        replaceData("Report Coordinator", 18, 11, ws, df)


    wb.save(pathName)


#Changes fields from Yes/No to a boolean

def fieldsToBoolean(columnName, targetCol):

    #Read the File, workbook, and Worksheet

    df = pd.read_excel("TestArcherSOXReportValidation.xlsx")

    wb = load_workbook("TestArcherSOXReportValidation.xlsx")

    ws = wb.active

    ws.title = "Sheet1"


    #Replace the values

    df[columnName].replace(regex = True, inplace = True, to_replace = "No", value = "false")

    df[columnName].replace(regex = True, inplace = True, to_replace = "", value = "na")

    df[columnName].replace(regex = True, inplace = True, to_replace = "Yes", value = "true")


    #Output to current excel doc

    df.to_excel('TestArcherSOXReportValidation.xlsx', sheet_name = "Sheet1", index = False, header = True)


def sortCollibraNames(pathName):


    nameColumnIndex = []


    if pathName == "Business Process":

        fileName = "TestCollibraBusinessProcessValidation.xlsx"

        nameColumnIndex = ["SOX Business Process Participants"]


    if pathName == "Controls":

        fileName = "TestCollibraControlValidation.xlsx"

        nameColumnIndex = ["SOX Control Owner", "SOX Point of Contact"]


    df = pd.read_excel(fileName)

    df.replace(np.nan, "", regex = True, inplace = True)


    for i in range(len(nameColumnIndex)):

        temp = []

        for row in range(df.shape[0]):


            temp = df.at[row,nameColumnIndex[i]].split("; ")


            temp.sort()


            df.at[row, nameColumnIndex[i]] = "; ".join(temp)

    df.to_excel(fileName, index = False)



#----- Business Process Validation -----#

def businessProcessValidation():

    #Set file path

    CBPpath = "C:\\Users\\JMurray64\\Documents\\Archer Integration Testing\\Integration Test 4\\TestCollibraBusinessProcessValidation.xlsx"


    #Place collibra column names in array

    CBP_column_names = {'Name': 'Process Name', 'SOX Business Process Owner': 'Business Process Owner', 'SOX Business Process Participants': 'Business Process Participants'}


    #Read the excel file dataset

    collibraBP = pd.read_excel(CBPpath)


    #Rename columns

    collibraBP.rename(columns=CBP_column_names, inplace = True)


    #Replace all not a NaN values with a blank

    collibraBP.replace(np.nan, '', regex=True, inplace = True)


    #Remove string leading and trailing spaces

    collibraBP.replace("\xa0", " ", regex = True, inplace = True)

    collibraBP.replace("\n", " ", regex = True, inplace = True)

    collibraBP["Definition"] = collibraBP["Definition"].str.strip()

    collibraBP["Definition"] = collibraBP["Definition"].str.rstrip("\n")


    #Makes sure Collibra names are formatted properly using shape[0] to just get the number of rows for the range

    for r in range(collibraBP.shape[0]):


        #As it loops, set the name to each row in Business Process Participants

        name = collibraBP.at[r, 'Business Process Participants']


        #Continue if it is a string

        if type(name) == str:


            #Continue if there is a new line within the text

            if '\n' in name:


                #**Split the name string into parts by each line into an array and sort.

                CBP_name_list = name.split('\n')

                CBP_name_list.sort()


                #**Join the names back together separated by a new line

                collibraBP.at[r, 'Business Process Participants'] = '\n'.join(CBP_name_list)

            else:

                continue

        else:

            continue


    #collibraBP.to_excel('Temp_Businesss_Process.xlsx', index = False, header = True)

    #Set file path

    ABPpath = "C:\\Users\\JMurray64\\Documents\\Archer Integration Testing\\Integration Test 4\\TestArcherBusinessProcessValidation.xlsx"


    #Read the excel file dataset

    archerBP = pd.read_excel(ABPpath)


    #Replace all not a number values with a blank

    archerBP.replace(np.nan, '', regex=True, inplace = True)


    #Remove string leading and trailing spaces

    archerBP["Description"] = archerBP["Description"].str.strip()


    #Loops through each row of the archer business process dataset

    for r in range(archerBP.shape[0]):


        #As it loops, set the name to each row in Business Process Participants

        name = archerBP.at[r, 'Business Process Participants']


        #Proceed if the data type of name is a string

        if type(name) == str:


            #Proceed if there is a new line in name

            if '\n' in name:


                #Split the name string into parts by each line into an array and sort.

                ABP_name_list = name.split('\n')

                ABP_name_list.sort()


                #Join the names back together separated by a new line

                archerBP.at[r, 'Business Process Participants'] = '\n'.join(ABP_name_list)

            else:

                continue

        else:

            continue


    #Set compareBP as boolean of T/F depending on if the datasets are a perfect match (Does this store an array of T/Fs?)

    compareBP = collibraBP.values == archerBP.values


    #**Bring out all row/columns where there are no matches between archer and collibra

    rows,cols = np.where(compareBP==False)


    #Loop for each row/col

    for item in zip(rows,cols):

        #locate each item for each non-matching row/column value in the collibra dataset and replace with "collibra value --> archer value"

        collibraBP.iloc[item[0], item[1]] = 'Archer --> {} \n\nCollibra --> {}'.format(archerBP.iloc[item[0], item[1]],collibraBP.iloc[item[0], item[1]])


    #writes the dataset to a new workbook

    collibraBP.to_excel('Collibra_to_Archer_BP_Validation.xlsx', index = False, header = True)


    #load workbook for openpyxl

    wb = load_workbook('Collibra_to_Archer_BP_Validation.xlsx')

    ws = wb.active

    rowIndex = 2


    #Create red color

    redFill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')


    #iterate through rows to highlight non-matching cells red

    for row in ws.iter_rows(min_row = 2, max_col = 5, max_row = 34, values_only = True):

        colIndex = 1

        for item in row:

            #fill each cell item meeting the conditions with red

            if item != None and "-->" in item:

                ws.cell(row = rowIndex, column = colIndex).fill = redFill

            colIndex += 1


        rowIndex += 1

    #Save the workbook

    wb.save('Collibra_to_Archer_BP_Validation.xlsx')

#----- END Business Process Validation -----#


#----- Control Validation -----#

def controlValidation():

    #Declare dictionary to change the header names

    CCV_headers = {'Full Name': 'Control ID', 'Name': 'Control Name', 'SOX Life Cycle Status': 'Life Cycle Status', 'enforces [Business Process] > Name': 'Business Process', 'SOX Control Owner': 'Control Owner', 'Last Updated in Aurora': 'Last Aurora Update', 'SOX Control Level': 'Control Level', 'SOX Control Purpose': 'Control Purpose', 'SOX Risk Rank': 'Risk Rank', 'SOX Point of Contact': 'Point of Contact', 'SOX Tier': 'Tier', 'SOX Control Rationale': 'Control Rationale', 'SOX Functional Group Owner': 'Functional Group Owner', 'SOX Functional Group Details': 'Functional Group Details', 'SOX Control Category': 'Control Category', 'SOX Audit Requirement': 'Audit Requirement', 'SOX Automation Status': 'Automation Status'}


    #Declare file path and read document to pandas dataframe

    CCVpath = "C:\\Users\\JMurray64\\Documents\\Archer Integration Testing\\Integration Test 4\\TestCollibraControlValidation.xlsx"

    collibraCV = pd.read_excel(CCVpath)


    #Rename columns using above dictionary

    collibraCV.rename(columns=CCV_headers, inplace = True)


    #Replace all NaN values with blanks

    collibraCV.replace(np.nan, '', regex=True, inplace = True)

    collibraCV.replace("\xa0", " ", regex = True, inplace = True)

    collibraCV.replace("\n ", " ", regex = True, inplace = True)

    collibraCV.replace("\n", " ", regex = True, inplace = True)

    collibraCV.replace("   ", " ", regex = True, inplace = True)

    collibraCV.replace("  ", " ", regex = True, inplace = True)

    collibraCV["Description"] = collibraCV["Description"].str.rstrip("\n")

    collibraCV["Description"] = collibraCV["Description"].str.strip()

    collibraCV["Control Name"] = collibraCV["Control Name"].str.rstrip("\n")

    collibraCV["Control Name"] = collibraCV["Control Name"].str.strip()

    collibraCV["Control Rationale"] = collibraCV["Control Rationale"].str.rstrip("\n")

    collibraCV["Control Rationale"] = collibraCV["Control Rationale"].str.strip()


    #Drop Control Procedure and Last updated in Aurora date

    collibraCV.drop(['SOX Control Procedure', 'Last Aurora Update'], axis=1, inplace=True) #delete unneccesary col


    #loop for the row length of the datafram

    for r in range(collibraCV.shape[0]):

        control_name = collibraCV.at[r, 'Control Name']

        #Splits out the control from the control name

        if ' - ' in control_name:

            dash = control_name.find(" - ")

            if "(" in control_name:

                lParenthesis = control_name.find("(")

                rParenthesis = control_name.find(")")

                if lParenthesis < dash and rParenthesis > dash:

                    dash = control_name.find(" - ",dash + 1)


            collibraCV.at[r, 'Control Name'] = control_name[dash+3:]

        else:

            continue


    #Declare dictionary to change the header names

    ACV_headers = {'Procedure ID': 'Control ID', 'Procedure Name': 'Control Name', 'Control Lifecycle Status': 'Life Cycle Status','Managing Group': 'SOX Managing Group', 'Last Updated Date': 'Last Aurora Update'}


    #Declare file path and read document to pandas dataframe

    ACVpath = "C:\\Users\\JMurray64\\Documents\\Archer Integration Testing\\Integration Test 4\\TestArcherControlValidation.xlsx"

    archerCV = pd.read_excel(ACVpath)

    #Rename columns using above dictionary

    archerCV.rename(columns = ACV_headers, inplace = True)

    #Replace all NaN values with blanks

    archerCV.replace(np.nan, '', regex=True, inplace = True)

    #Drop Last Aurora Update column

    archerCV.drop('Last Aurora Update', axis=1, inplace=True)

    archerCV.replace("\xa0", " ", regex = True, inplace = True)

    archerCV.replace("\n ", " ", regex = True, inplace = True)

    archerCV.replace("\n", " ", regex = True, inplace = True)

    archerCV.replace("   ", " ", regex = True, inplace = True)

    archerCV.replace("  ", " ", regex = True, inplace = True)

    archerCV["Description"] = archerCV["Description"].str.strip()

    archerCV["Control Rationale"] = archerCV["Control Rationale"].str.strip()

    archerCV["Control Name"] = archerCV["Control Name"].str.strip()

    archerCV["Control ID"] = archerCV["Control ID"].str.strip()

 

    #Merge the dataframes on select columns to find the records in common

    combined = pd.merge(collibraCV, archerCV, on = ['Control ID', 'Control Name', 'Description', 'Life Cycle Status', 'Business Process', 'Control Owner', 'SOX Managing Group', 'Control Level', 'Control Purpose', 'Risk Rank', 'Point of Contact', 'Tier', 'Control Rationale','Functional Group Owner', 'Functional Group Details', 'Control Category','Audit Requirement', 'Automation Status'], how = 'inner')


    #Stack the combined data onto the archer dataframe

    arch = pd.concat([archerCV, combined], sort = False)

    #Remove duplicate values from the archer + combined dataframe to get only rows without matches

    arch.drop_duplicates(keep=False, inplace = True)

    #Identify the source of the data as Archer

    arch['Source'] = 'Archer'

    #Stack the combined data onto the collibra dataframe

    collib = pd.concat([collibraCV, combined], sort = False)

    #Remove duplicate values from the archer + combined dataframe to get only rows without matches

    collib.drop_duplicates(keep=False, inplace = True)

    #Identify the source of the data as Collibra

    collib['Source'] = 'Collibra'


    #Stack the rows that dont match from archer and collibra together and sort on Control ID

    merged = pd.concat([arch, collib], sort = False)

    merged.sort_values(['Control ID'], inplace = True)


    merged.reset_index(inplace=True, drop = True)

    header_listCV = list(merged)

    control_IDs = set(merged['Control ID'].to_list())

    #merged.set_index(keys='Control ID',inplace=True)

    try:

        for n in range(merged.shape[0]+2):

            for header in header_listCV[1:-1]:

                if merged.at[n, 'Control ID'] == merged.at[n+1, 'Control ID']:

                    if merged.at[n, header] == merged.at[n+1, header]:

                        merged.at[n+1, header] = 'MATCH'

                    else:

                        continue

                else:

                    continue

    except KeyError:

        pass


    #Output the resulting dataset tot an excel doc

    merged.to_excel('Collibra_to_Archer_CV_Validation.xlsx', index = True, header = True)


#----- END Control Validation -----#

def newLineToColon(columnName, df):

    #Loops through each item item in the dataset

    for i in df.index:


        if df[columnName][i] != "" and df[columnName][i] != None:


            #Splits each person into their own list item

            nameList = str(df[columnName][i]).split("\n")


            if columnName == "Data Source Application" or columnName == "Reporting Tool Application":

                #Sort list and join to one string 

                names = ";".join(nameList)

            else:

                names = "; ".join(nameList)


            #Replace oringal entry with newly modified string

            df[columnName][i] = names


#----- Report Validation -----#

def reportValidation():

    #Set file path and read document to panda dataframe

    CRVpath = "C:\\Users\\JMurray64\\Documents\\Archer Integration Testing\\Integration Test 4\\TestCollibraSOXReportValidation.xlsx"

    collibraRV = pd.read_excel(CRVpath)


    #Drop the Aliases Column

    collibraRV.drop('Aliases', axis=1, inplace=True)


    #Declare dictionary to change the header names

    CRV_headers = {'Full Name': 'Report Tracking ID', 'Name': 'Report Name', 'SOX Report Purpose': 'Report Purpose', 'SOX Life Cycle Status': 'Report Lifecycle Status', 'Aurora Key Fields': 'Key Fields', 'Aurora Data Source Application': 'Data Source Application', 'Aurora Reporting Tool Application': 'Reporting Tool Application', 'SOX Managing Group': 'Managing Group', 'governed by [Control] > Name': 'Control Procedures', 'SOX Report Coordinator': 'Report Coordinator', 'SOX Key Report': 'Key?', 'SOX Key Rationale': 'Key Rationale', 'SOX Report Type': 'Report Type', 'SOX Report Sub-Type': 'Report Sub-Type', 'Impacts Financial Statement?': 'Report Impacts Financial Statement?', 'Impacts Financial Statement Comment': 'Report Impacts Financial Statement Comment', 'Last Updated in Aurora': 'Last Updated Date'}


    #Rename columns using above dictionary

    collibraRV.rename(columns = CRV_headers, inplace=True)

    collibraRV.replace(np.nan, '', regex=True, inplace = True)

    collibraRV.replace("\xa0", " ", regex = True, inplace = True)

    collibraRV.replace("\n ", "\n", regex = True, inplace = True)

    collibraRV["Key Fields"].replace("\n", " ", regex = True, inplace = True)

    collibraRV["Report Name"].replace("\n", " ", regex = True, inplace = True)

    collibraRV.replace("   ", " ", regex = True, inplace = True)

    collibraRV.replace("  ", " ", regex = True, inplace = True)


    collibraRV["Key Fields"] = collibraRV["Key Fields"].str.strip("\n")

    collibraRV["Key Fields"] = collibraRV["Key Fields"].str.strip()

    collibraRV["Report Purpose"] = collibraRV["Report Purpose"].str.strip("\n")

    collibraRV["Report Purpose"] = collibraRV["Report Purpose"].str.strip()

    collibraRV["Key Rationale"] = collibraRV["Key Rationale"].str.strip("\n")

    collibraRV["Key Rationale"] = collibraRV["Key Rationale"].str.strip()

    collibraRV["Report Impacts Financial Statement Comment"] = collibraRV["Report Impacts Financial Statement Comment"].str.strip("\n")

    collibraRV["Report Impacts Financial Statement Comment"] = collibraRV["Report Impacts Financial Statement Comment"].str.strip()


    #Set file path and read document to panda dataframe

    ARVpath = "C:\\Users\\JMurray64\\Documents\\Archer Integration Testing\\Integration Test 4\\TestArcherSOXReportValidation.xlsx"

    archerRV = pd.read_excel(ARVpath)


    #Replace all NaN values with blanks

    archerRV.replace(np.nan, '', regex=True, inplace = True)


    #Declare dictionary to change the header names

    ARV_headers = {'Step 1a) Control Procedures Association': 'Control Procedures'}


    #Drop the Additonal Report Names and Last Updated Date columns

    archerRV.drop(['Additional Report Names', 'Last Updated Date'], axis=1, inplace=True)


    #Rename columns using above dictionary

    archerRV.rename(columns = ARV_headers, inplace=True)

    archerRV["Key Fields"] = archerRV["Key Fields"].str.strip()

    archerRV["Report Purpose"] = archerRV["Report Purpose"].str.strip()

    archerRV["Report Purpose"] = archerRV["Report Purpose"].str.strip("\n")

    archerRV["Report Name"] = archerRV["Report Name"].str.strip()

    archerRV["Report Impacts Financial Statement Comment"] = archerRV["Report Impacts Financial Statement Comment"].str.strip()

    #archerRV["Control Procedure"] = archerRV["Control Procedure"].str.strip()


    archerRV.replace("\t", " ", regex = True, inplace = True)

    archerRV.replace("\n ", "\n", regex = True, inplace = True)

    archerRV["Report Name"].replace("\n", " ", regex = True, inplace = True)

    archerRV["Key Fields"].replace("\n", " ", regex = True, inplace = True)

    archerRV.replace("   ", " ", regex = True, inplace = True)

    archerRV.replace("  ", " ", regex = True, inplace = True)

    archerRV.replace("<W>", "", regex = True, inplace = True)


    archerRV["Key Fields"] = archerRV["Key Fields"].str.strip()

    archerRV["Report Purpose"] = archerRV["Report Purpose"].str.strip()

    archerRV["Report Purpose"] = archerRV["Report Purpose"].str.strip("\n")

    archerRV["Report Name"] = archerRV["Report Name"].str.strip()

    archerRV["Report Impacts Financial Statement Comment"] = archerRV["Report Impacts Financial Statement Comment"].str.strip()

    archerRV["Control Procedures"] = archerRV["Control Procedures"].str.strip()


    newLineToColon("Data Source Application", archerRV)

    newLineToColon("Reporting Tool Application", archerRV)

    newLineToColon("Key Rationale", archerRV)


    #Replace T/F values with Y/N in he Key and Report Impacts Financial Statment columns

    collibraRV["Key?"]= collibraRV["Key?"].replace(True, "Yes")

    collibraRV["Key?"]= collibraRV["Key?"].replace(False, "No")

    collibraRV["Report Impacts Financial Statement?"]= collibraRV["Report Impacts Financial Statement?"].replace(True, "Yes")

    collibraRV["Report Impacts Financial Statement?"]= collibraRV["Report Impacts Financial Statement?"].replace(False, "No")


    collibraRV.drop('Report Impacts Financial Statement?', axis=1, inplace=True)

    collibraRV.drop('Report Impacts Financial Statement Comment', axis=1, inplace=True)

    archerRV.drop('Report Impacts Financial Statement?', axis=1, inplace=True)

    archerRV.drop('Report Impacts Financial Statement Comment', axis=1, inplace=True)




    #loop for the row length of the datafram

    collibraRV.drop(['SOX Control Procedure','Last Updated Date'], axis=1, inplace=True)


    #Split out the control ID from the text

    for n in range(collibraRV.shape[0]):

            control_name = collibraRV.at[n, 'Control Procedures']

            if " - " in control_name:

                dash = control_name.find("-")

                if "(" in control_name:

                    lParenthesis = control_name.find("(")

                    rParenthesis = control_name.find(")")

                    if lParenthesis < dash and rParenthesis > dash:

                        dash = control_name.find(" - ", dash+2) + 1


            collibraRV.at[n, 'Control Procedures'] = control_name[:dash - 1]



    #Merge the dataframes on select columns to find the records in common

    combinedRV = pd.merge(collibraRV, archerRV, on = ['Report Tracking ID', 'Report Name', 'Report Purpose', 'Report Lifecycle Status', 'Key Fields', 'Data Source Application', 'Reporting Tool Application', 'Managing Group', 'Control Procedures','Report Coordinator', 'Key?', 'Key Rationale', 'Report Type', 'Report Sub-Type'], how = 'inner')


    #Stack the combined data onto the archer dataframe

    archRV = pd.concat([archerRV, combinedRV], sort = False)


    #Remove duplicate values from the archer + combined data frame to get only rows without matches

    archRV.drop_duplicates(keep=False, inplace = True)


    #Identify the source of the data as Archer

    archRV['Source'] = 'Archer'


    #Stack the combined data onto the collibra dataframe

    collibRV = pd.concat([collibraRV, combinedRV], sort = False)


    #Remove duplicate values from the collibra + combined data frame to get only rows without matches

    collibRV.drop_duplicates(keep=False, inplace = True)


    #Identify the source of the data as Collibra

    collibRV['Source'] = 'Collibra'


    #Stack the rows that dont match from archer and collibra together and sort on Control ID

    mergedRV = pd.concat([archRV, collibRV], sort = False)

    mergedRV.sort_values(['Report Tracking ID', 'Control Procedures'], inplace = True)



    mergedRV.reset_index(inplace=True, drop = True)

    header_listRV = list(mergedRV)


    try:

        for n in range(mergedRV.shape[0]+2):

            for header in header_listRV[1:-1]:

                if mergedRV.at[n, 'Report Tracking ID'] == mergedRV.at[n+1, 'Report Tracking ID']:

                    if mergedRV.at[n, header] == mergedRV.at[n+1, header]:

                        mergedRV.at[n+1, header] = 'MATCH'

                    else:

                        continue

                else:

                    continue

    except (ValueError, KeyError):

        pass



    #Output the resulting dataset tot an excel doc

    mergedRV.to_excel('Collibra_to_Archer_RV_Validation.xlsx', index = False, header = True)


removeFormatText("TestCollibraSOXReportValidation.xlsx")

removeFormatText("TestCollibraBusinessProcessValidation.xlsx")

removeFormatText("TestCollibraControlValidation.xlsx")


reformatNames("Archer Business Process")

reformatNames("Archer Controls")

reformatNames("Archer Reports")


sortCollibraNames("Business Process")

sortCollibraNames("Controls")

parseSOXControls("TestArcherSOXReportValidation.xlsx")

businessProcessValidation()

controlValidation()

reportValidation()

